"""Single entrypoint for transport results dashboard generation.

This file is a thin wrapper around `results_analysis.leap_series_analysis_workflow`.
Running this workflow calls that module's `build_config(...)` and `run_with_config(...)`
to perform comparison, chart generation, and per-economy dashboard build.
"""
#%%
from __future__ import annotations

import importlib.util
import math
import os
from pathlib import Path
import re
import sys
from typing import Any

import pandas as pd
from functions.transport_branch_paths import extract_transport_branch_tuple

REPO_ROOT = Path(__file__).resolve().parents[2]
if REPO_ROOT.exists() and str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
if (REPO_ROOT / "codebase").exists() and str(REPO_ROOT / "codebase") not in sys.path:
    sys.path.insert(0, str(REPO_ROOT / "codebase"))

from results_analysis.leap_series_analysis_workflow import _build_sheet_dashboards
from results_analysis.leap_series_analysis_workflow import build_config, run_with_config
from results_analysis.leap_series_comparison import _write_transport_comparison_charts
from results_analysis.stock_projection_exploration import (
    ExplorationConfig,
    run_exploration,
)


# Notebook/script-editable defaults
SCENARIOS = ("Reference", "Target")
INCLUDE_ECONOMIES: tuple[str, ...] | None = None
INPUT_DIR = REPO_ROOT / "results/checkpoint_audit"
OUTPUT_DIR = REPO_ROOT / "results/diagnostics/transport_results_series_comparison"
INCLUDE_STOCK_PROXIES = True
STOCK_PROXY_DIR = REPO_ROOT / "results/diagnostics/stock_projection_exploration"
INCLUDE_APEC_AGGREGATE = False
INCLUDE_LEAP_RESULTS_COMPARISON = True
LEAP_UTILITIES_REPO = Path("C:/Users/Work/github/leap_utilities/")  # Update this path to your local leap_utilities repository
if os.getcwd() != REPO_ROOT:
    print(
        f"[WARN] Current working directory is not repo root; "
        f"relative paths may not resolve as expected: {os.getcwd()}"
    )
LEAP_RESULTS_TABLES_DIR = LEAP_UTILITIES_REPO / "data/leap results tables"
CHART_BACKEND = "plotly"  # "static" or "plotly"
# Set to None to keep default series behavior.
# Allowed categories:
# - "input"
# - "pre"
# - "reconciled"
# - "reconciled_plus_alt"
# - "checkpoint_direct_proxy"
# - "sales_flow_projected_proxy"
SERIES_CATEGORIES: tuple[str, ...] | None = (
    "input",
    "pre",
    "reconciled",
    "reconciled_plus_alt",
    "checkpoint_direct_proxy",
    "sales_flow_projected_proxy",
)


def _normalize_filename_token(value: object) -> str:
    return re.sub(r"[\s_]+", "", str(value or "").strip().lower())


def _safe_filename_token(value: object) -> str:
    text = str(value).strip() if value is not None else ""
    if not text:
        return "series"
    safe = "".join(ch if ch.isalnum() or ch in {"_", "-"} else "_" for ch in text)
    return safe.strip("_") or "series"


def _parse_region_from_workbook_metadata(
    workbook_path: Path,
    scenario: str,
) -> str | None:
    try:
        import openpyxl
    except Exception as exc:
        print(f"[WARN] openpyxl unavailable for region discovery ({exc}); skipping: {workbook_path.name}")
        return None

    scenario_key = str(scenario).strip().lower()
    pattern = re.compile(r"Scenario:\s*(.+?)(?:,\s*Region:\s*(.+?))?(?:,\s*All Fuels)?\s*$", flags=re.IGNORECASE)
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"[WARN] Failed to open workbook for region discovery ({workbook_path}): {exc}")
        return None

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            a2 = ws["A2"].value
            if not a2:
                continue
            text = str(a2).strip()
            if not text:
                continue
            match = pattern.search(text)
            if not match:
                continue
            parsed_scenario = str(match.group(1) or "").strip().lower()
            parsed_region = str(match.group(2) or "").strip()
            if parsed_scenario == scenario_key and parsed_region:
                return parsed_region
    finally:
        wb.close()
    return None


def _discover_economies_for_scenarios(
    *,
    input_dir: str | Path,
    scenarios: tuple[str, ...],
) -> set[str]:
    root = Path(input_dir).expanduser()
    if not root.is_absolute():
        root = (REPO_ROOT / root).resolve()
    if not root.exists():
        return set()

    economies: set[str] = set()
    for scenario in scenarios:
        pattern = re.compile(
            rf"^transport_pre_recon_vs_raw_disaggregated_(.+)_{re.escape(str(scenario))}\.csv$",
            flags=re.IGNORECASE,
        )
        for path in root.glob(f"transport_pre_recon_vs_raw_disaggregated_*_{scenario}.csv"):
            match = pattern.match(path.name)
            if match:
                economies.add(str(match.group(1)).strip())
    return economies


def _run_or_load_leap_results_comparisons(
    *,
    scenarios: tuple[str, ...],
    include_economies: tuple[str, ...] | None,
    input_dir: str | Path,
    output_dir: str | Path,
) -> list[dict[str, Any]]:
    print("[INFO] LEAP results comparison mode enabled; scanning optional workbook inputs.")
    leap_input_dir = LEAP_RESULTS_TABLES_DIR.resolve()
    if not leap_input_dir.exists():
        print(f"[WARN] LEAP results tables directory missing; skipping integration: {leap_input_dir}")
        return []

    all_workbooks = sorted(
        [
            *leap_input_dir.glob("*.xlsx"),
            *leap_input_dir.glob("*.xlsm"),
            *leap_input_dir.glob("*.xls"),
        ]
    )
    all_workbooks = [wb for wb in all_workbooks if not wb.name.startswith("~$")]
    if not all_workbooks:
        print(f"[WARN] No LEAP workbooks found in {leap_input_dir}; skipping integration.")
        return []
    print(f"[INFO] Found {len(all_workbooks)} workbook(s) in {leap_input_dir}")

    requested_economies = (
        {str(e).strip() for e in include_economies if str(e).strip()}
        if include_economies
        else _discover_economies_for_scenarios(input_dir=input_dir, scenarios=scenarios)
    )
    requested_economies.discard("00_APEC")
    if not requested_economies:
        print("[WARN] No requested economies resolved for LEAP results comparison; skipping integration.")
        return []

    external_module_path = LEAP_UTILITIES_REPO / "codebase/functions/leap_series_comparison.py"
    if not external_module_path.exists():
        print(f"[WARN] External comparison module missing; skipping integration: {external_module_path}")
        return []

    utilities_repo = LEAP_UTILITIES_REPO.resolve()
    if str(utilities_repo) not in sys.path:
        sys.path.insert(0, str(utilities_repo))
    spec = importlib.util.spec_from_file_location("leap_utilities_series_comparison", external_module_path)
    if spec is None or spec.loader is None:
        print(f"[WARN] Unable to load external comparison module; skipping integration: {external_module_path}")
        return []

    external_mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = external_mod
    spec.loader.exec_module(external_mod)
    cfg_cls = getattr(external_mod, "TransportResultsComparisonConfig", None)
    run_fn = getattr(external_mod, "run_transport_results_table_comparison", None)
    if cfg_cls is None or run_fn is None:
        print("[WARN] External comparison config/function missing; skipping integration.")
        return []

    integration_root = Path(output_dir).expanduser()
    if not integration_root.is_absolute():
        integration_root = (REPO_ROOT / integration_root).resolve()
    integration_root = integration_root / "leap_results_tables_comparison"

    comparison_runs: list[dict[str, Any]] = []
    normalized_files = {wb: _normalize_filename_token(wb.name) for wb in all_workbooks}
    for economy in sorted(requested_economies):
        econ_token = _normalize_filename_token(economy)
        for scenario in scenarios:
            scenario_token = _normalize_filename_token(scenario)
            candidates = [
                wb
                for wb, norm_name in normalized_files.items()
                if econ_token in norm_name and scenario_token in norm_name
            ]
            if not candidates:
                print(
                    f"[WARN] No LEAP workbook filename matched economy/scenario "
                    f"tokens for {economy} / {scenario}; skipping."
                )
                continue

            candidates = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
            selected = candidates[0]
            if len(candidates) > 1:
                print(
                    f"[WARN] Multiple workbook matches for {economy} / {scenario}; "
                    f"using newest: {selected.name}"
                )
            selected_norm = _normalize_filename_token(selected.name)
            if econ_token not in selected_norm or scenario_token not in selected_norm:
                print(
                    f"[WARN] Selected workbook failed token validation for {economy} / {scenario}; "
                    f"skipping: {selected.name}"
                )
                continue

            region = _parse_region_from_workbook_metadata(selected, str(scenario))
            if not region:
                print(
                    f"[WARN] Could not infer region from workbook metadata for "
                    f"{selected.name} [{scenario}]; skipping."
                )
                continue

            pair_output_dir = integration_root / f"{_safe_filename_token(economy)}__{_safe_filename_token(scenario)}"
            comparison_long_csv = pair_output_dir / "comparison_long.csv"
            charts_dir = pair_output_dir / "charts"
            if comparison_long_csv.exists() and charts_dir.exists():
                print(
                    f"[INFO] Reusing existing LEAP comparison artifacts for {economy}/{scenario}: "
                    f"{pair_output_dir}"
                )
                comparison_runs.append(
                    {
                        "economy": economy,
                        "scenario": str(scenario),
                        "workbook": selected,
                        "comparison_long_csv": comparison_long_csv,
                        "charts_dir": charts_dir,
                    }
                )
                continue

            print(
                f"[INFO] Running LEAP results comparison for {economy}/{scenario} "
                f"using {selected.name}"
            )
            try:
                cfg = cfg_cls(
                    leap_results_file=selected,
                    economy=economy,
                    scenario=str(scenario),
                    region=region,
                    branch_sector_mapping_csv=LEAP_UTILITIES_REPO / "config/leap_transport_branch_to_ninth_sector_map.csv",
                    fuel_aliases_csv=LEAP_UTILITIES_REPO / "config/leap_transport_fuel_aliases.csv",
                    code_to_name_path=LEAP_UTILITIES_REPO / "config/sector_fuel_codes_to_names.xlsx",
                    code_to_name_sheet="code_to_name",
                    esto_data_path=LEAP_UTILITIES_REPO / "data/00APEC_2024_low.csv",
                    ninth_data_path=LEAP_UTILITIES_REPO / "data/merged_file_energy_ALL_20250814_pre_trump.csv",
                    subtotal_mapping_path=LEAP_UTILITIES_REPO / "config/ESTO_subtotal_mapping.xlsx",
                    ninth_to_esto_mapping_path=LEAP_UTILITIES_REPO / "config/ninth_pairs_to_esto_pairs.xlsx",
                    output_dir=pair_output_dir,
                )
                artifacts = run_fn(cfg)
            except Exception as exc:
                print(
                    f"[WARN] LEAP results comparison failed for {economy}/{scenario}; "
                    f"skipping integration for this pair: {exc}"
                )
                continue

            artifacts_csv = Path(artifacts.comparison_long_csv)
            artifacts_charts = Path(artifacts.charts_dir)
            if not artifacts_csv.exists() or not artifacts_charts.exists():
                print(
                    f"[WARN] LEAP comparison artifacts missing for {economy}/{scenario}; "
                    "skipping integration for this pair."
                )
                continue
            comparison_runs.append(
                {
                    "economy": economy,
                    "scenario": str(scenario),
                    "workbook": selected,
                    "comparison_long_csv": artifacts_csv,
                    "charts_dir": artifacts_charts,
                }
            )

    if not comparison_runs:
        print("[WARN] No LEAP comparison runs/artifacts available for dashboard integration.")
    else:
        print(f"[INFO] LEAP comparison runs/artifacts ready: {len(comparison_runs)}")
    return comparison_runs


def _inject_leap_results_comparison_into_dashboard(
    *,
    base_comparison_long_csv: Path,
    base_charts_dir: Path,
    comparison_runs: list[dict[str, Any]],
) -> int:
    if not base_comparison_long_csv.exists():
        print(f"[WARN] Base comparison_long.csv missing; cannot inject LEAP artifacts: {base_comparison_long_csv}")
        return 0
    if not base_charts_dir.exists():
        print(f"[WARN] Base charts_dir missing; cannot inject LEAP artifacts: {base_charts_dir}")
        return 0
    if not comparison_runs:
        print("[INFO] No LEAP comparison artifacts to inject.")
        return 0

    base_df = pd.read_csv(base_comparison_long_csv)
    injected_frames: list[pd.DataFrame] = []
    injected_rows = 0
    major_transport_type_map = {
        "passenger road": "Passenger road",
        "freight road": "Freight road",
        "passenger non road": "Passenger non-road",
        "freight non road": "Freight non-road",
        "pipeline transport": "Pipelines",
    }

    def _map_branch_to_major_transport_type(branch_path: object) -> str:
        logical_tuple = extract_transport_branch_tuple(str(branch_path or ""))
        if not logical_tuple:
            return "LEAP"
        top_level = str(logical_tuple[0]).strip()
        key = top_level.lower()
        if key in major_transport_type_map:
            return major_transport_type_map[key]
        if key == "international transport":
            return "International transport"
        tail = str(logical_tuple[-1]).strip()
        key = tail.lower()
        if key in major_transport_type_map:
            return major_transport_type_map[key]
        if key == "international transport":
            return "International transport"
        if key == "demand":
            return "LEAP demand total"
        return tail

    def _map_fuel_label(value: object) -> str:
        token = str(value or "").strip()
        if token.lower() in {"__total__", "total"}:
            return "Total"
        return token

    def _auto_scale_leap_values(leap_df: pd.DataFrame, base_df_local: pd.DataFrame) -> pd.DataFrame:
        if leap_df.empty or base_df_local.empty:
            return leap_df
        key_cols = ["economy", "scenario", "metric", "major_transport_type", "fuel_label", "year"]
        required = set(key_cols + ["input_value"])
        if not required.issubset(leap_df.columns):
            return leap_df
        if not set(key_cols).issubset(base_df_local.columns):
            return leap_df

        base = base_df_local.copy()
        base["base_value"] = (
            pd.to_numeric(base.get("reconciled_plus_alt_value"), errors="coerce")
            .combine_first(pd.to_numeric(base.get("reconciled_value"), errors="coerce"))
            .combine_first(pd.to_numeric(base.get("pre_value"), errors="coerce"))
            .combine_first(pd.to_numeric(base.get("input_value"), errors="coerce"))
        )
        base = base[key_cols + ["base_value"]].dropna(subset=["base_value"])
        if base.empty:
            return leap_df

        leap = leap_df.copy()
        leap["scenario"] = leap["scenario"].astype(str).str.replace(r"\s+LEAP$", "", regex=True)
        leap["input_value"] = pd.to_numeric(leap["input_value"], errors="coerce")
        probe = leap[key_cols + ["input_value"]].rename(columns={"input_value": "leap_value"})
        merged = probe.merge(base, on=key_cols, how="inner")
        merged["ratio"] = pd.to_numeric(merged["leap_value"], errors="coerce") / pd.to_numeric(
            merged["base_value"], errors="coerce"
        )
        merged = merged.replace([float("inf"), float("-inf")], pd.NA).dropna(subset=["ratio"])
        merged = merged[merged["ratio"].abs() > 0]
        if merged.empty:
            return leap_df

        ratio_median = float(pd.to_numeric(merged["ratio"].abs(), errors="coerce").median())
        if not math.isfinite(ratio_median) or ratio_median <= 0:
            return leap_df

        # Prefer stable engineering-scale factors to avoid overfitting noisy overlap matches.
        candidate_factors = [1.0, 1e-3, 1e-6, 1e3]
        factor = min(candidate_factors, key=lambda f: abs(math.log10(max(ratio_median * f, 1e-30))))
        if abs(math.log10(max(ratio_median * factor, 1e-30))) > 1.2:
            return leap_df

        out = leap_df.copy()
        for col in ("input_value", "pre_value", "reconciled_value", "reconciled_plus_alt_value"):
            if col in out.columns:
                out[col] = pd.to_numeric(out[col], errors="coerce") * factor
        print(
            "[INFO] Applied LEAP auto-scale factor "
            f"({factor:g}) using overlap median ratio ~{ratio_median:.4g}."
        )
        return out

    for run in comparison_runs:
        run_csv = Path(run["comparison_long_csv"])
        economy = str(run["economy"])
        if not run_csv.exists():
            print(f"[WARN] Missing optional LEAP comparison CSV; skipping: {run_csv}")
            continue

        try:
            leap_df = pd.read_csv(run_csv)
        except Exception as exc:
            print(f"[WARN] Failed to read optional LEAP comparison CSV ({run_csv}): {exc}")
            continue

        required_cols = {"branch_path", "fuel_label", "year", "scenario", "leap_value", "reference_value"}
        if not required_cols.issubset(leap_df.columns):
            print(f"[WARN] Optional LEAP CSV missing required columns; skipping: {run_csv}")
            continue

        leap_df["economy"] = economy
        leap_df["metric"] = "energy"
        leap_df["major_transport_type"] = leap_df["branch_path"].map(_map_branch_to_major_transport_type)
        leap_df["fuel_label"] = leap_df["fuel_label"].map(_map_fuel_label)
        leap_df["scenario"] = leap_df["scenario"].map(lambda value: f"{str(value).strip()} LEAP")
        leap_df["year"] = pd.to_numeric(leap_df["year"], errors="coerce")
        leap_df["input_value"] = pd.to_numeric(leap_df["leap_value"], errors="coerce")
        leap_df["pre_value"] = pd.NA
        leap_df["reconciled_value"] = pd.NA
        leap_df["reconciled_plus_alt_value"] = pd.NA
        if "unit_label" in base_df.columns:
            leap_df["unit_label"] = "PJ"
        if "scale_factor" in base_df.columns:
            leap_df["scale_factor"] = 1e9
        leap_df = _auto_scale_leap_values(leap_df, base_df)

        aligned = pd.DataFrame(index=leap_df.index, columns=base_df.columns)
        for col in aligned.columns:
            if col in leap_df.columns:
                aligned[col] = leap_df[col]

        required_group_cols = [
            c for c in ("economy", "metric", "major_transport_type", "fuel_label") if c in aligned.columns
        ]
        if required_group_cols:
            aligned = aligned.dropna(subset=required_group_cols, how="any")
        if aligned.empty:
            continue

        injected_frames.append(aligned)
        injected_rows += len(aligned)

    if not injected_frames:
        print("[INFO] No optional LEAP rows injected into base comparison table.")
        return 0

    injected_df = pd.concat(injected_frames, ignore_index=True)
    merged_df = pd.concat([base_df, injected_df], ignore_index=True)
    dedupe_cols = [
        c
        for c in ["economy", "scenario", "metric", "major_transport_type", "fuel_label", "year"]
        if c in merged_df.columns
    ]
    if dedupe_cols:
        merged_df = merged_df.drop_duplicates(subset=dedupe_cols, keep="last")
    merged_df.to_csv(base_comparison_long_csv, index=False)

    print(
        "[INFO] Injected optional LEAP comparison rows into dashboard source: "
        f"rows_added={injected_rows}"
    )
    return injected_rows


def _add_nonroad_aggregates_for_leap_overlay(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    if df.empty:
        return df, 0
    if "scenario" not in df.columns or "major_transport_type" not in df.columns:
        return df, 0

    leap_present = df["scenario"].astype(str).str.contains(r"\sLEAP$", regex=True, na=False).any()
    if not leap_present:
        return df, 0

    value_cols = [
        col
        for col in (
            "input_value",
            "pre_value",
            "reconciled_value",
            "reconciled_plus_alt_value",
            "checkpoint_direct_proxy_value",
            "sales_flow_projected_proxy_value",
        )
        if col in df.columns
    ]
    if not value_cols:
        return df, 0

    metric_sum = {"activity", "energy", "stock"}
    metric_factor = {"efficiency", "intensity", "mileage"}
    passenger_components = ["Passenger Air", "Passenger Rail", "Passenger Ship"]
    freight_components = ["Freight Air", "Freight Rail", "Freight Ship"]
    component_to_group = {
        **{k: "Passenger non-road" for k in passenger_components},
        **{k: "Freight non-road" for k in freight_components},
    }

    working = df.copy()
    working["metric"] = working["metric"].astype(str)
    working["major_transport_type"] = working["major_transport_type"].astype(str)
    non_leap = working[~working["scenario"].astype(str).str.contains(r"\sLEAP$", regex=True, na=False)].copy()
    non_leap = non_leap[non_leap["major_transport_type"].isin(component_to_group.keys())].copy()
    if non_leap.empty:
        print("[INFO] No base Air/Rail/Ship rows available for non-road aggregate synthesis.")
        return df, 0

    index_keys = [
        k
        for k in ("economy", "scenario", "year", "fuel_label", "major_transport_type")
        if k in non_leap.columns
    ]
    if len(index_keys) < 5:
        return df, 0

    for col in value_cols:
        non_leap[col] = pd.to_numeric(non_leap[col], errors="coerce")

    activity_lookup = non_leap[non_leap["metric"].str.lower() == "activity"][
        index_keys + value_cols
    ].rename(columns={c: f"__activity_{c}" for c in value_cols})
    stock_lookup = non_leap[non_leap["metric"].str.lower() == "stock"][
        index_keys + value_cols
    ].rename(columns={c: f"__stock_{c}" for c in value_cols})

    synth_rows: list[pd.DataFrame] = []
    for metric_name, metric_df in non_leap.groupby(non_leap["metric"].str.lower(), dropna=False):
        if metric_name not in metric_sum and metric_name not in metric_factor:
            continue
        metric_df = metric_df.copy()
        metric_df["major_transport_type"] = metric_df["major_transport_type"].map(component_to_group)
        group_keys = [
            k for k in ("economy", "scenario", "year", "fuel_label", "major_transport_type", "metric") if k in metric_df.columns
        ]
        if metric_name in metric_sum:
            agg = metric_df.groupby(group_keys, dropna=False)[value_cols].sum(min_count=1).reset_index()
            synth_rows.append(agg)
            continue

        base_keys = [k for k in ("economy", "scenario", "year", "fuel_label", "major_transport_type") if k in metric_df.columns]
        weight_source = activity_lookup
        weight_prefix = "__activity_"
        if metric_name == "mileage":
            weight_source = stock_lookup
            weight_prefix = "__stock_"

        merged = metric_df.merge(weight_source, on=base_keys, how="left")
        out_groups: list[dict[str, Any]] = []
        for key_vals, sub in merged.groupby(group_keys, dropna=False):
            row: dict[str, Any] = {}
            if isinstance(key_vals, tuple):
                for idx, key_name in enumerate(group_keys):
                    row[key_name] = key_vals[idx]
            else:
                row[group_keys[0]] = key_vals

            for col in value_cols:
                vals = pd.to_numeric(sub[col], errors="coerce")
                wcol = f"{weight_prefix}{col}"
                weights = pd.to_numeric(sub.get(wcol), errors="coerce") if wcol in sub.columns else pd.Series(pd.NA, index=sub.index, dtype="float64")
                valid = vals.notna() & weights.notna() & weights.gt(0)
                if valid.any():
                    row[col] = float((vals[valid] * weights[valid]).sum() / weights[valid].sum())
                else:
                    row[col] = float(vals.dropna().mean()) if vals.notna().any() else pd.NA
            out_groups.append(row)
        synth_rows.append(pd.DataFrame(out_groups))

    if not synth_rows:
        return df, 0

    synthetic = pd.concat([f for f in synth_rows if not f.empty], ignore_index=True)
    if synthetic.empty:
        return df, 0

    aligned = pd.DataFrame(index=synthetic.index, columns=working.columns)
    for col in aligned.columns:
        if col in synthetic.columns:
            aligned[col] = synthetic[col]

    merged_out = pd.concat([working, aligned], ignore_index=True)
    dedupe_cols = [
        c
        for c in ("economy", "scenario", "metric", "major_transport_type", "fuel_label", "year")
        if c in merged_out.columns
    ]
    if dedupe_cols:
        merged_out = merged_out.drop_duplicates(subset=dedupe_cols, keep="last")
    print(
        "[INFO] Added synthetic non-road base rows for LEAP overlay: "
        f"{len(aligned)} row(s) using Air/Rail/Ship aggregation."
    )
    return merged_out, int(len(aligned))


def _ensure_stock_proxy_files(
    *,
    scenarios: tuple[str, ...],
    include_economies: tuple[str, ...] | None,
    input_dir: str | Path,
    stock_proxy_dir: str | Path,
) -> None:
    requested_economies = (
        {str(e).strip() for e in include_economies if str(e).strip()}
        if include_economies
        else _discover_economies_for_scenarios(input_dir=input_dir, scenarios=scenarios)
    )
    # APEC aggregate is not used for economy-level dashboard overlays.
    requested_economies.discard("00_APEC")

    if not requested_economies:
        print("[INFO] No economies discovered for stock proxy generation; skipping.")
        return

    output_dir = Path(stock_proxy_dir).expanduser()
    if not output_dir.is_absolute():
        output_dir = (REPO_ROOT / output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    for economy in sorted(requested_economies):
        try:
            cfg = ExplorationConfig(
                economy=economy,
                scenarios=tuple(str(s).strip() for s in scenarios if str(s).strip()),
                output_dir=output_dir,
            )
            written = run_exploration(cfg)
            print(
                "[INFO] Stock proxy generation complete "
                f"for {economy}: wrote {len(written)} file(s)."
            )
        except Exception as exc:
            print(
                "[WARN] Stock proxy generation failed "
                f"for {economy}; continuing without overlays for this economy: {exc}"
            )


def run_dashboard_workflow(
    *,
    scenarios: tuple[str, ...] = SCENARIOS,
    include_economies: tuple[str, ...] | None = INCLUDE_ECONOMIES,
    input_dir: str | Path = INPUT_DIR,
    output_dir: str | Path = OUTPUT_DIR,
    include_stock_proxies: bool = INCLUDE_STOCK_PROXIES,
    stock_proxy_dir: str | Path = STOCK_PROXY_DIR,
    include_apec_aggregate: bool = INCLUDE_APEC_AGGREGATE,
    include_leap_results_comparison: bool = INCLUDE_LEAP_RESULTS_COMPARISON,
    chart_backend: str = CHART_BACKEND,
    series_categories: tuple[str, ...] | None = SERIES_CATEGORIES,
):
    if include_stock_proxies:
        _ensure_stock_proxy_files(
            scenarios=scenarios,
            include_economies=include_economies,
            input_dir=input_dir,
            stock_proxy_dir=stock_proxy_dir,
        )

    # Delegate all core analysis/dashboard logic to leap_series_analysis_workflow.
    cfg = build_config(
        scenarios=scenarios,
        include_economies=include_economies,
        input_dir=input_dir,
        output_dir=output_dir,
        include_stock_proxies=include_stock_proxies,
        stock_proxy_dir=stock_proxy_dir,
        include_apec_aggregate=include_apec_aggregate,
        chart_backend=chart_backend,
        series_categories=series_categories,
    )
    artifacts = run_with_config(cfg)

    if not include_leap_results_comparison:
        print("[INFO] LEAP results comparison mode disabled.")
        return artifacts

    comparison_runs = _run_or_load_leap_results_comparisons(
        scenarios=scenarios,
        include_economies=include_economies,
        input_dir=input_dir,
        output_dir=output_dir,
    )
    injected_rows = _inject_leap_results_comparison_into_dashboard(
        base_comparison_long_csv=Path(artifacts.comparison_long_csv),
        base_charts_dir=Path(artifacts.charts_dir),
        comparison_runs=comparison_runs,
    )

    if injected_rows > 0:
        try:
            merged_df = pd.read_csv(Path(artifacts.comparison_long_csv))
            merged_df, added_nonroad_rows = _add_nonroad_aggregates_for_leap_overlay(merged_df)
            if added_nonroad_rows > 0:
                merged_df.to_csv(Path(artifacts.comparison_long_csv), index=False)
            _write_transport_comparison_charts(
                merged_df,
                Path(artifacts.charts_dir),
                chart_backend=chart_backend,
                series_categories=series_categories,
            )
            print("[INFO] Regenerated base charts after LEAP series injection.")
        except Exception as exc:
            print(f"[WARN] Failed to regenerate charts after LEAP series injection: {exc}")

        dashboard_index = _build_sheet_dashboards(
            output_dir=Path(artifacts.comparison_long_csv).parent,
            comparison_long_csv=Path(artifacts.comparison_long_csv),
            charts_dir=Path(artifacts.charts_dir),
        )
        if dashboard_index:
            print(f"[INFO] Rebuilt dashboards with optional LEAP comparison integration: {dashboard_index}")
    else:
        print("[INFO] Optional LEAP comparison integration skipped (no injectible artifacts).")

    return artifacts


if __name__ == "__main__":
    run_dashboard_workflow()
#%%
